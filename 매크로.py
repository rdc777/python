from ctypes.wintypes import WCHAR
import tkinter.messagebox as msgbox
import pyautogui as pag
import time
import sys 
import keyword
import keyboard
from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np


root = Tk()
root.title("매크로")
root.geometry("640x380") # 가로 * 세로

def new_file():
    listbox.delete(0,END)
# 파일 불러오기
def open_file():
    
    path = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("excel files", "*.xlsx") ,
                                          ("all files", "*.*")))
    
    df = pd.read_excel(path, header = None)
    #df = [x for x in tuple((df[0] ,df[1]))]
    # print(df)
    #df = df.squeeze()
    # print(df)
    # 결측값(nan 값 변환)
    df = df.fillna("")
    # for d in range(len(df)): 
    i0_t = [i0 for i0 in df[0]]
    i1_t = [i1 for i1 in df[1]]
    #리스트로 변환 (A,B 열만)

    # print(i0_t)
    # print(i1_t)
    #     df_t = []
    #     df_t.append(i0_t, i1_t)
    # 튜플로 변경하고 좌표 와 그외 명령어 구분
    for i in range(len(i0_t)):
        if i1_t[i] == "":
            t = (i0_t[i])
            print(t)
            listbox.insert(i,t)
        else:
            t = (i0_t[i],i1_t[i])
            print(t)
            listbox.insert(i,t)

    # print(t)
    # print(df_t)    
    # for i in range(len(t)):
    #     if t[i] == 1:
    #         listbox.insert(i,t[i])

    # print(path)

def save_file():
    #저장 위치 지정
    save_c = filedialog.asksaveasfilename(initialdir="/", title="Select file",
                                          filetypes=(("excel files", "*.xlsx") ,
                                          ("all files", "*.*")))

    wb = Workbook() # 새 워크북 생성
    ws = wb.active # 현재 활성화된 sheet 가져옴
    ws.title = "m" # sheet 의 이름을 변경
    save_l = []

    for lb in listbox.get(0,END):    
        save_l.append(lb)
    #리스트 박스값을 리스트에 저장

    # save_l = [x_l for x_l in listbox.get(0,END)]
    #좌표의 경우 2열을 사용 하여 하기 때문에 구분
    a = 0
    for l in save_l:
        if len(l) == 2:
            a =a+1
            ws["A"+ str(a)] =l[0]
            ws["B"+ str(a)] = l[1]
            print(l)
            a = int(a)
        else:
            a =a+1
            ws["A"+ str(a)] = l
            print(l)
            a = int(a)
    
    wb.save(save_c+".xlsx")
    
    wb.close()
   # print(save_c)


menu = Menu(root)
menu_file = Menu(menu, tearoff=0)
menu_file.add_command(label="New File",command=new_file)#리스트박스 초기화
menu_file.add_command(label="New Window")#장식
menu_file.add_separator()
menu_file.add_command(label="Open",command=open_file)#파일 불러오기
menu_file.add_separator()
menu_file.add_command(label="Save",command=save_file)#매크로 저장하여 보관
#, state="disable") # 비활성화
menu_file.add_separator()
menu_file.add_command(label="Exit", command=root.quit)#종료
menu.add_cascade(label="File", menu=menu_file)

list_frame = Frame(root)
list_frame.pack(fill="both", padx=5, pady=5)

scrollbar = Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")

listbox = Listbox(list_frame, selectmode="extended", height=15, yscrollcommand=scrollbar.set)
listbox.pack(side="left", fill="both", expand=True)
scrollbar.config(command=listbox.yview)

def btncmd():
    
    b =0
    if listbox.size() ==0:
        for pos_l in pos_lits():
            listbox.insert(b, pos_l)
            b = b + 1
    else: 

        for pos_l in pos_lits():
            tupB = (b+1,)
            tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
            listbox.insert(tupSum,pos_l)
            b = b + 1

# 좌표에 대하여 리스트에 저장한 값을 반환

def pos_lits(pos_list=[]):
#pos_list = []
    i = 0
    pos_list.clear()
    while i ==0:
        while True:
            if keyboard.is_pressed('F3'):
                #pos_list = []
                pos = pag.position()
                print(pos)
                pos_list.append(pos)
                break;

        while True:
            if keyboard.is_pressed('F4'):
            
                break;
            elif keyboard.is_pressed('F5'):
                i=+1
                return pos_list
                break;   
                    
        while i==1:
            break;

btn = Button(root,padx=5, pady=5, width=8, text="좌표", command=btncmd)
btn.pack(side="left")

def btncmd1():
    det = listbox.curselection()   
    for d in reversed(det):#reversed 리스트의 원소를 거꾸로 뒤집어 반환
        listbox.delete(d)
        #print(d)
   

btn = Button(root,padx=5, pady=5, width=8, text="삭제", command=btncmd1)
btn.pack(side="left")

#리스트작성
def btncmd_c():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,"Ctrl+c")

def btncmd_v():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,"Ctrl+v")

def btncmd_cl():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,"click")
def time1():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,"Time")


ctrlc = Button(root,padx=5, pady=5, width=5, text="Ctrl+C", command=btncmd_c)
ctrlc.pack(side="right")
ctrlv = Button(root,padx=5, pady=5, width=5, text="Ctrl+V", command=btncmd_v)
ctrlv.pack(side="right")
click1 = Button(root,padx=5, pady=5, width=5, text="Click", command=btncmd_cl)
click1.pack(side="right")
times = Button(root,padx=5, pady=5, width=5, text="Time", command=time1)
times.pack(side="right")



def play():
    for play_1 in listbox.get(0,END):
        if len(play_1) == 2:
            pos_x, pos_y = play_1
            pag.moveTo(pos_x, pos_y)
        # elif len(play_1) ==3:   
        #     pag.keyDown('ctrl') # ctrl 키를 누른 상태를 유지합니다.
        #     pag.press('c') # c key를 입력합니다. 
        #     pag.keyUp('ctrl') # ctrl 키를 뗍니다. 
        elif len(play_1) ==6:
            pag.keyDown('ctrl') # ctrl 키를 누른 상태를 유지합니다.
            pag.press(str(play_1[5:])) # c/v key를 입력합니다. 
            pag.keyUp('ctrl') # ctrl 키를 뗍니다.
        elif len(play_1) ==5:
            pag.click()
        elif len(play_1) ==4:
            time.sleep(1)  

play = Button(root,padx=5, pady=5, width=8, text="Play", command=play)
play.pack(side="left")


def x_lable():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,x.get()+'x')
    
def y_lable():
    tupB = (1,)
    tupSum=tuple(sum(elem) for elem in zip(listbox.curselection(), tupB))
    listbox.insert(tupSum,x.get()+'y')

def l_lable():
    req = int(l.get())
    for l_l in range(0,req):
        j = 0
        for play_2 in listbox.get(0,END):
            j = j + 1
            if len(play_2) == 2:
                pos_x, pos_y = play_2
                pag.moveTo(pos_x, pos_y)
            elif len(play_2) ==6:
                pag.keyDown('ctrl') # ctrl 키를 누른 상태를 유지합니다.
                pag.press(str(play_2[5:])) # c/v key를 입력합니다. 
                pag.keyUp('ctrl') # ctrl 키를 뗍니다.
            elif len(play_2) ==5:
                pag.click()
            elif len(play_2) ==4:
                time.sleep(1)
            elif str(play_2[-1:])== 'x':
                if len(listbox.get(j-2,)) ==2:    
                    tupB = (int(play_2[0:-1]),0)
                    # print(tupB)
                    print(play_2)
                    # print(listbox.get(j-2,))# 이전값 반환
                    tupSum=tuple(sum(elem) for elem in zip(listbox.get(j-2,), tupB))#변경될 값 좌표 지정
                    listbox.insert(j-2,tupSum)#리스트에 값 반환
                    listbox.delete(j-1)
                    # print(j-1)
                else:
                    tupB = (int(play_2[0:-1]),0)
                    tupSum=tuple(sum(elem) for elem in zip(listbox.get(j-3,), tupB))#변경될 값 좌표 지정
                    listbox.insert(j-3,tupSum)#리스트에 값 반환
                    listbox.delete(j-2)
                    
            elif str(play_2[-1:])== 'y':
                if len(listbox.get(j-2,)) ==2:
                    tupB = (0,int(play_2[0:-1]))
                    # print(tupB)
                    # print(play_2)
                    # print(listbox.get(j-2,))# 이전값 반환
                    tupSum=tuple(sum(elem) for elem in zip(listbox.get(j-2,), tupB))#변경될 값 좌표 지정
                    listbox.insert(j-2,tupSum)#리스트에 값 반환
                    listbox.delete(j-1)
                    # print(j-1)
                else:
                    tupB = (0,int(play_2[0:-1]))
                    tupSum=tuple(sum(elem) for elem in zip(listbox.get(j-3,), tupB))#변경될 값 좌표 지정
                    listbox.insert(j-3,tupSum)#리스트에 값 반환
                    listbox.delete(j-2)
    msgbox.showinfo("알림", str(l_l+1)+ "건 완료되었습니다.")

x = Entry(root, width=4)
x.pack(side="left")
x.insert(0, "x축")
btnx = Button(root, text="클릭", command=x_lable)
btnx.pack(side="left")

y = Entry(root, width=4)
y.pack(side="left")
y.insert(0, "y축")
btny = Button(root, text="클릭", command=y_lable)
btny.pack(side="left")

l = Entry(root, width=4)
l.pack(side="left")
l.insert(0, "반복")
btny = Button(root, text="클릭", command=l_lable)
btny.pack(side="left")

root.resizable(True, False)

root.config(menu=menu)

root.mainloop()

#pyinstaller -w -F C:\file_python\set1\매크로.py
#pyinstaller -w C:\file_python\set1\매크로.py